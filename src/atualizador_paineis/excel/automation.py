from __future__ import annotations

import logging
import shutil
import warnings
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from atualizador_paineis.core.config import PanelConfig
from atualizador_paineis.core.errors import ExcelAutomationError

LOGGER = logging.getLogger(__name__)

PACKAGE_COMPONENTS = {
    "worksheets": "xl/worksheets/sheet",
    "charts": "xl/charts/chart",
    "pivot_tables": "xl/pivotTables/pivotTable",
    "slicer_caches": "xl/slicerCaches/slicerCache",
    "slicers": "xl/slicers/slicer",
    "drawings": "xl/drawings/drawing",
}


@dataclass(frozen=True, slots=True)
class ColumnUpdate:
    column_name: str
    values: tuple[Any, ...]


@dataclass(frozen=True, slots=True)
class TableUpdate:
    sheet_name: str
    table_name: str
    rows_to_delete: tuple[int, ...]
    remaining_row_count: int
    appended_data: pd.DataFrame
    column_updates: tuple[ColumnUpdate, ...] = ()


@dataclass(frozen=True, slots=True)
class CellUpdate:
    sheet_name: str
    table_name: str
    row_index: int
    column_name: str
    value: Any


def ensure_panel_is_closed(panel_path: Path) -> None:
    lock_file = panel_path.with_name(f"~${panel_path.name}")
    if lock_file.exists():
        raise ExcelAutomationError(
            "O painel está aberto no Excel. Feche o arquivo e tente novamente."
        )


def workbook_structure(path: Path) -> dict[str, int]:
    with ZipFile(path) as package:
        names = package.namelist()
    return {
        component: sum(
            name.startswith(prefix) and name.endswith(".xml") and "/_rels/" not in name
            for name in names
        )
        for component, prefix in PACKAGE_COMPONENTS.items()
    }


def read_table_dataframe(path: Path, sheet_name: str, table_name: str) -> pd.DataFrame:
    """Lê somente os dados de uma tabela estruturada sem salvar a planilha."""
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            workbook = load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:
        raise ExcelAutomationError(f"Não foi possível ler o histórico do painel: {exc}") from exc
    try:
        if sheet_name not in workbook.sheetnames:
            raise ExcelAutomationError(f"A aba '{sheet_name}' não existe no painel.")
        worksheet = workbook[sheet_name]
        if table_name not in worksheet.tables:
            raise ExcelAutomationError(
                f"A tabela '{table_name}' não foi encontrada na aba '{sheet_name}'."
            )
        min_col, min_row, max_col, max_row = range_boundaries(
            worksheet.tables[table_name].ref
        )
        headers = [
            worksheet.cell(min_row, column).value for column in range(min_col, max_col + 1)
        ]
        rows = [
            [worksheet.cell(row, column).value for column in range(min_col, max_col + 1)]
            for row in range(min_row + 1, max_row + 1)
        ]
        return pd.DataFrame(rows, columns=headers)
    finally:
        workbook.close()


def _excel_value(value: Any) -> Any:
    if value is None or value is pd.NA:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, (datetime, date, str, int, float, bool)):
        return value
    if hasattr(value, "item"):
        return value.item()
    return str(value)


def _disconnect_query_table(worksheet: Any, table: Any) -> Any:
    try:
        query_table = table.QueryTable
    except Exception:
        return table
    if query_table is None:
        return table

    table_name = table.Name
    table_style = table.TableStyle
    table_range = table.Range.Address
    LOGGER.info("Convertendo a tabela %s em tabela local", table_name)
    table.Unlist()
    recreated = worksheet.ListObjects.Add(1, worksheet.Range(table_range), None, 1)
    recreated.Name = table_name
    if table_style:
        recreated.TableStyle = table_style
    return recreated


def _contiguous_groups(indices: tuple[int, ...]) -> list[tuple[int, int]]:
    if not indices:
        return []
    groups: list[tuple[int, int]] = []
    start = previous = indices[0]
    for index in indices[1:]:
        if index == previous + 1:
            previous = index
            continue
        groups.append((start, previous))
        start = previous = index
    groups.append((start, previous))
    return groups


def _apply_table_update(workbook: Any, update: TableUpdate) -> None:
    worksheet = workbook.Worksheets(update.sheet_name)
    try:
        table = worksheet.ListObjects(update.table_name)
    except Exception as exc:
        raise ExcelAutomationError(
            f"A tabela '{update.table_name}' não foi encontrada na aba '{update.sheet_name}'."
        ) from exc

    table = _disconnect_query_table(worksheet, table)
    start_row = table.Range.Row
    start_column = table.Range.Column
    column_count = len(update.appended_data.columns)

    for first, last in reversed(_contiguous_groups(update.rows_to_delete)):
        worksheet.Range(
            worksheet.Cells(start_row + 1 + first, start_column),
            worksheet.Cells(start_row + 1 + last, start_column + column_count - 1),
        ).Delete()

    remaining_count = table.ListRows.Count
    if remaining_count != update.remaining_row_count:
        raise ExcelAutomationError(
            f"A tabela '{update.table_name}' ficou com {remaining_count} linhas após a "
            f"substituição; eram esperadas {update.remaining_row_count}."
        )

    for column_update in update.column_updates:
        if len(column_update.values) != remaining_count:
            raise ExcelAutomationError(
                f"A atualização da coluna '{column_update.column_name}' possui "
                f"{len(column_update.values)} valores; eram esperados {remaining_count}."
            )
        if not remaining_count:
            continue
        column_index = list(update.appended_data.columns).index(column_update.column_name)
        column_number = start_column + column_index
        values = tuple((_excel_value(value),) for value in column_update.values)
        worksheet.Range(
            worksheet.Cells(start_row + 1, column_number),
            worksheet.Cells(start_row + remaining_count, column_number),
        ).Value = values

    appended_count = len(update.appended_data)
    total_count = remaining_count + appended_count
    new_range = worksheet.Range(
        worksheet.Cells(start_row, start_column),
        worksheet.Cells(start_row + total_count, start_column + column_count - 1),
    )
    table.Resize(new_range)

    chunk_size = 5_000
    for offset in range(0, appended_count, chunk_size):
        chunk = update.appended_data.iloc[offset : offset + chunk_size]
        values = tuple(
            tuple(_excel_value(value) for value in row)
            for row in chunk.itertuples(index=False, name=None)
        )
        first_row = start_row + 1 + remaining_count + offset
        worksheet.Range(
            worksheet.Cells(first_row, start_column),
            worksheet.Cells(first_row + len(values) - 1, start_column + column_count - 1),
        ).Value = values


def _apply_cell_update(workbook: Any, update: CellUpdate) -> None:
    worksheet = workbook.Worksheets(update.sheet_name)
    try:
        table = worksheet.ListObjects(update.table_name)
        column = table.ListColumns(update.column_name)
        data_body = column.DataBodyRange
    except Exception as exc:
        raise ExcelAutomationError(
            f"A coluna '{update.column_name}' não foi encontrada na tabela "
            f"'{update.table_name}' da aba '{update.sheet_name}'."
        ) from exc

    if update.row_index < 0 or update.row_index >= table.ListRows.Count:
        raise ExcelAutomationError(
            f"A linha {update.row_index + 1} não existe na tabela '{update.table_name}'."
        )
    data_body.Cells(update.row_index + 1, 1).Value = _excel_value(update.value)


def _refresh_pivots(workbook: Any) -> None:
    errors: list[str] = []
    for index in range(1, workbook.PivotCaches().Count + 1):
        try:
            cache = workbook.PivotCaches().Item(index)
            cache.MissingItemsLimit = 0
            cache.Refresh()
        except Exception as exc:
            errors.append(f"cache {index}: {exc}")
    for worksheet in workbook.Worksheets:
        try:
            pivot_tables = worksheet.PivotTables()
            for index in range(1, pivot_tables.Count + 1):
                pivot_tables.Item(index).RefreshTable()
        except Exception as exc:
            errors.append(f"aba {worksheet.Name}: {exc}")
    if errors:
        raise ExcelAutomationError(
            "Não foi possível atualizar todas as tabelas dinâmicas: " + "; ".join(errors)
        )


def update_staged_workbook(
    staged_path: Path,
    updates: tuple[TableUpdate, ...],
    cell_updates: tuple[CellUpdate, ...] = (),
) -> None:
    try:
        import pythoncom
        import win32com.client
        from win32com.client import gencache
    except ImportError as exc:
        raise ExcelAutomationError(
            "O componente de automação do Excel não está instalado. Reinstale o aplicativo."
        ) from exc

    excel = None
    workbook = None
    pythoncom.CoInitialize()
    try:
        excel = gencache.EnsureDispatch(win32com.client.DispatchEx("Excel.Application"))
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        workbook = excel.Workbooks.Open(str(staged_path.resolve()), UpdateLinks=0, ReadOnly=False)
        for update in updates:
            _apply_table_update(workbook, update)
        _refresh_pivots(workbook)
        for update in cell_updates:
            _apply_cell_update(workbook, update)
        workbook.Save()
        workbook.Close(SaveChanges=True)
        workbook = None
    except ExcelAutomationError:
        raise
    except Exception as exc:
        raise ExcelAutomationError(f"Falha ao atualizar o painel pelo Excel: {exc}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def validate_staged_workbook(
    staged_path: Path,
    config: PanelConfig,
    expected_raw_rows: int,
    expected_treated_rows: int,
    expected_structure: dict[str, int],
) -> None:
    actual_structure = workbook_structure(staged_path)
    if actual_structure != expected_structure:
        differences = ", ".join(
            f"{key}: {expected_structure[key]} → {actual_structure.get(key, 0)}"
            for key in expected_structure
            if expected_structure[key] != actual_structure.get(key, 0)
        )
        raise ExcelAutomationError(
            "A estrutura visual do painel foi alterada durante a atualização: " + differences
        )
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(staged_path, read_only=False, data_only=False)
    try:
        for sheet_name, table_name, expected_rows in (
            (config.raw_sheet, config.raw_table, expected_raw_rows),
            (config.treated_sheet, config.treated_table, expected_treated_rows),
        ):
            if sheet_name not in workbook.sheetnames:
                raise ExcelAutomationError(
                    f"A aba '{sheet_name}' desapareceu do painel atualizado."
                )
            worksheet = workbook[sheet_name]
            if table_name not in worksheet.tables:
                raise ExcelAutomationError(
                    f"A tabela '{table_name}' desapareceu do painel atualizado."
                )
            table_ref = worksheet.tables[table_name].ref
            min_col, min_row, max_col, max_row = range_boundaries(table_ref)
            del min_col, max_col
            actual_rows = max_row - min_row
            if actual_rows != expected_rows:
                raise ExcelAutomationError(
                    f"A tabela '{table_name}' deveria ter {expected_rows} registros, "
                    f"mas possui {actual_rows}."
                )
    finally:
        workbook.close()


def publish_panel(
    original_path: Path,
    output_path: Path,
    backup_dir: Path,
    staging_dir: Path,
    config: PanelConfig,
    updates: tuple[TableUpdate, ...],
    expected_raw_rows: int,
    expected_treated_rows: int,
    cell_updates: tuple[CellUpdate, ...] = (),
) -> Path:
    ensure_panel_is_closed(original_path)
    if output_path != original_path and output_path.exists():
        ensure_panel_is_closed(output_path)
        raise ExcelAutomationError(
            f"Já existe outro painel de destino chamado '{output_path.name}'."
        )

    backup_dir.mkdir(parents=True, exist_ok=True)
    staging_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    staged_path = staging_dir / f"painel-{timestamp}.xlsx"
    backup_path = backup_dir / f"{original_path.stem}-{timestamp}{original_path.suffix}"
    shutil.copy2(original_path, staged_path)
    expected_structure = workbook_structure(original_path)

    try:
        update_staged_workbook(staged_path, updates, cell_updates)
        validate_staged_workbook(
            staged_path,
            config,
            expected_raw_rows,
            expected_treated_rows,
            expected_structure,
        )
        original_path.replace(backup_path)
        try:
            staged_path.replace(output_path)
        except Exception:
            shutil.copy2(backup_path, original_path)
            raise
        return backup_path
    finally:
        if staged_path.exists():
            staged_path.unlink()
